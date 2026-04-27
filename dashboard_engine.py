from __future__ import annotations

import json
import math
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from calendar import monthrange
from pathlib import Path
from openpyxl.utils.cell import range_boundaries
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel

CODE_RE = re.compile(r"^\d+(?:\.\d+)*\.?$")

STATUS_ALIASES = {
    "contract prep": "contract prep.",
    "contract preparation": "contract prep.",
    "pct approval": "PCT/approval",
    "pct/approval": "PCT/approval",
    "closed/finalized": "finalized",
    "close/finalized": "finalized",
}

STATUS_PROGRESS = {
    None: 0,
    "planned": 8,
    "enquiry": 25,
    "negotiation": 50,
    "contract prep.": 75,
    "PCT/approval": 90,
    "closed": 100,
    "finalized": 100,
    "backup": 0,
}

STATUS_ORDER = [
    "planned",
    "enquiry",
    "negotiation",
    "contract prep.",
    "PCT/approval",
    "closed",
    "finalized",
]

DEFAULT_CONFIG = {
    "project_title": "Waste-to-Energy Project Dashboard",
    "executive_note": "Vista executive focalizzata su budget, procurement e stato ordini.",
    "currency": "PLN",
    "budget": {
        "sheet_name": "Włocławek Budget",
        "summary_row": 12,
        "total_column": "AB",
        "detail_start_row": 13,
        "detail_end_row": 279,
        "detail_exec_column": "W",
        "detail_total_column": "AB",
        "pl_desc_column": "C",
        "en_desc_column": "D",
        "contractor_column": "AC",
        "order_no_column": "AD",
        "curve_label": "TKW Razem",
        "curve_label_column": "C",
        "curve_header_row": 11,
        "curve_fallback_row": 428,
        "curve_start_column": "AM",
        "curve_end_column": "BZ"
    },
    "procurement": {
        "sheet_name": "Procurement Schedule",
        "summary_row": 16,
        "updated_budget_column": "S",
        "contracted_column": "T",
        "legacy_budget_column": "R",
        "detail_start_row": 18,
        "status_column": "Q",
    },
    "portfolio_curve": {"sheet_name": "10288_ITPO PROJEKT-SCuvrve"},
    "ec_decision": {
        "sheet_name": "EC decision",
        "header_row": 3,
        "data_start_row": 4,
        "columns": ["D", "E", "G", "I", "J", "K", "L", "M"]
    },
    "completed_statuses": ["closed", "finalized", "PCT/approval"],
    "contract_prep_status": "contract prep.",
    "thresholds": {
        "variance_amber_pct": 2.0,
        "variance_red_pct": 6.0,
        "contract_coverage_amber_pct": 55.0,
        "contract_coverage_red_pct": 40.0,
    },
}


def col_to_index(col: str) -> int:
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column: {col}")
        n = n * 26 + ord(ch) - 64
    return n


def normalize_code(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        if float(value).is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if CODE_RE.match(s):
            return s.rstrip(".")
    return None


def safe_num(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    return 0.0


def first_non_empty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def normalize_status(status: Any) -> str | None:
    if status is None:
        return None
    s = str(status).strip()
    if not s:
        return None
    lower = s.lower()
    canonical = STATUS_ALIASES.get(lower, s)
    for known in STATUS_ORDER:
        if canonical.lower() == known.lower():
            return known
    return canonical


def to_iso(dt: Any) -> str | None:
    if isinstance(dt, datetime):
        return dt.date().isoformat()
    if isinstance(dt, date):
        return dt.isoformat()
    return None


def month_key(dt: datetime | date) -> str:
    return f"{dt.year:04d}-{dt.month:02d}"


def month_label(month: str) -> str:
    year, mon = month.split("-")
    names = ["gen", "feb", "mar", "apr", "mag", "giu", "lug", "ago", "set", "ott", "nov", "dic"]
    return f"{names[int(mon)-1]}-{year}"


def iter_months(start: str, end: str) -> list[str]:
    sy, sm = map(int, start.split("-"))
    ey, em = map(int, end.split("-"))
    out: list[str] = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        out.append(f"{y:04d}-{m:02d}")
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    return out


def tone_from_status(status: str | None) -> str:
    if status in {"finalized", "closed", "PCT/approval"}:
        return "green"
    if status in {"contract prep.", "negotiation"}:
        return "amber"
    if status in {"planned", "enquiry"}:
        return "blue"
    return "neutral"




def month_end(dt: datetime | date) -> datetime:
    return datetime(dt.year, dt.month, monthrange(dt.year, dt.month)[1])


def parse_month_value(value: Any) -> str | None:
    if value in (None, "", 0):
        return None
    if isinstance(value, datetime):
        return month_key(value)
    if isinstance(value, date):
        return month_key(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return month_key(dt)
        except Exception:
            pass
        if 1900 <= int(value) <= 2100 and float(value).is_integer():
            return f"{int(value):04d}-01"
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%m.%Y", "%Y-%m", "%m-%Y", "%m/%Y", "%Y/%m", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m")
            except Exception:
                pass
        if re.fullmatch(r"\d{4}", s):
            return f"{s}-01"
        if re.fullmatch(r"\d{2}\.\d{4}", s):
            return datetime.strptime(s, "%m.%Y").strftime("%Y-%m")
    return None


def parse_sheet_range(ref: str) -> tuple[str, str]:
    if '!' not in ref:
        raise ValueError(f"Invalid ref: {ref}")
    sheet, cells = ref.split('!', 1)
    return sheet.strip("'"), cells


def normalize_chart_label(value: Any) -> str:
    month = parse_month_value(value)
    if month:
        return month_label(month)
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return str(value).strip()


def to_float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    try:
        return float(str(value).strip().replace(',', '.'))
    except Exception:
        return None

@dataclass
class Node:
    code: str
    parent: str | None
    depth: int
    desc_pl: str | None = None
    desc_en: str | None = None
    buyer: str | None = None
    site_resp: str | None = None
    tme_resp: str | None = None
    category: str | None = None
    comment: str | None = None
    companies: set[str] = field(default_factory=set)
    statuses: list[str] = field(default_factory=list)
    direct_budget_ab: float = 0.0
    direct_budget_exec: float = 0.0
    direct_legacy_budget: float = 0.0
    direct_updated_budget: float = 0.0
    direct_contracted: float = 0.0
    direct_result: float = 0.0
    rows_budget: int = 0
    rows_proc: int = 0
    children: list[str] = field(default_factory=list)
    dates: dict[str, list[datetime]] = field(
        default_factory=lambda: {
            "start_proc": [],
            "deadline_proc": [],
            "deadline_closing": [],
            "start_exec": [],
            "end_exec": [],
        }
    )
    rollup: dict[str, Any] = field(default_factory=dict)


class DashboardEngine:
    def __init__(self, config_path: Path | None = None):
        self.config_path = config_path
        self.config = self.load_config(config_path)

    @staticmethod
    def load_config(config_path: Path | None) -> dict[str, Any]:
        config = json.loads(json.dumps(DEFAULT_CONFIG))
        if config_path and config_path.exists():
            incoming = json.loads(config_path.read_text(encoding="utf-8"))
            DashboardEngine._deep_update(config, incoming)
        return config

    @staticmethod
    def _deep_update(target: dict[str, Any], source: dict[str, Any]) -> None:
        for key, value in source.items():
            if isinstance(value, dict) and isinstance(target.get(key), dict):
                DashboardEngine._deep_update(target[key], value)
            else:
                target[key] = value

    def save_config(self, config_path: Path | None = None) -> None:
        path = config_path or self.config_path
        if not path:
            raise ValueError("No config path provided")
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(self.config, indent=2, ensure_ascii=False), encoding="utf-8")

    def read_budget_records(self, path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        cfg = self.config["budget"]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[cfg["sheet_name"]]

        summary = {
            "projectName": ws["C5"].value,
            "budgetExecTotal": safe_num(ws[f"W{cfg['summary_row']}"] .value),
            "budgetAbTotal": safe_num(ws[f"{cfg['total_column']}{cfg['summary_row']}"] .value),
        }

        pl_idx = col_to_index(cfg["pl_desc_column"]) - 1
        en_idx = col_to_index(cfg["en_desc_column"]) - 1
        exec_idx = col_to_index(cfg["detail_exec_column"]) - 1
        total_idx = col_to_index(cfg["detail_total_column"]) - 1
        contractor_idx = col_to_index(cfg["contractor_column"]) - 1
        order_idx = col_to_index(cfg["order_no_column"]) - 1

        records: list[dict[str, Any]] = []
        for idx, row in enumerate(
            ws.iter_rows(
                min_row=cfg["detail_start_row"],
                max_row=cfg["detail_end_row"],
                max_col=max(total_idx, contractor_idx, order_idx) + 1,
                values_only=True,
            ),
            start=cfg["detail_start_row"],
        ):
            code = normalize_code(row[0])
            if not code:
                continue
            records.append(
                {
                    "row": idx,
                    "code": code,
                    "desc_pl": row[pl_idx],
                    "desc_en": row[en_idx],
                    "exec_budget": row[exec_idx],
                    "budget_ab": row[total_idx],
                    "contractor": row[contractor_idx],
                    "order_no": row[order_idx],
                }
            )
        return records, summary

    def read_proc_records(self, path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        cfg = self.config["procurement"]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[cfg["sheet_name"]]

        updated_col = cfg["updated_budget_column"]
        contracted_col = cfg["contracted_column"]
        legacy_col = cfg["legacy_budget_column"]

        summary = {
            "updatedBudgetTotal": safe_num(ws[f"{updated_col}{cfg['summary_row']}"] .value),
            "contractedTotal": safe_num(ws[f"{contracted_col}{cfg['summary_row']}"] .value),
            "legacyBudgetTotal": safe_num(ws[f"{legacy_col}{cfg['summary_row']}"] .value),
        }

        records: list[dict[str, Any]] = []
        for idx, row in enumerate(ws.iter_rows(min_row=cfg["detail_start_row"], max_row=ws.max_row, max_col=23, values_only=True), start=cfg["detail_start_row"]):
            code = normalize_code(row[0])
            if not code:
                continue
            records.append(
                {
                    "row": idx,
                    "code": code,
                    "buyer": row[1],
                    "site_resp": row[2],
                    "tme_resp": row[3],
                    "desc_pl": row[4],
                    "desc_en": row[5],
                    "category": row[6],
                    "company": row[7],
                    "start_proc": row[11],
                    "deadline_proc": row[12],
                    "deadline_closing": row[13],
                    "start_exec": row[14],
                    "end_exec": row[15],
                    "status": normalize_status(row[16]),
                    "legacy_budget": row[17],
                    "updated_budget": row[18],
                    "contracted": row[19],
                    "result": row[20],
                    "other": row[21],
                    "comment": row[22],
                }
            )
        return records, summary


    def extract_budget_curve(self, budget_path: Path) -> dict[str, Any]:
        wb = openpyxl.load_workbook(budget_path, data_only=True, read_only=False)
        ws = wb[self.config["budget"]["sheet_name"]]
        label_col = col_to_index(self.config["budget"].get("curve_label_column", "C"))
        label_text = (self.config["budget"].get("curve_label") or "TKW Razem").strip().lower()
        start_col = col_to_index(self.config["budget"].get("curve_start_column", "AM"))
        end_col = col_to_index(self.config["budget"].get("curve_end_column", "BZ"))
        header_row = int(self.config["budget"].get("curve_header_row", 11))
        fallback_row = int(self.config["budget"].get("curve_fallback_row", 428))

        curve_row = None
        search_candidates = list(range(max(header_row + 1, fallback_row - 80), fallback_row + 81))
        for row_idx in search_candidates:
            raw = ws.cell(row_idx, label_col).value
            if raw and str(raw).strip().lower() == label_text:
                curve_row = row_idx
                break
        if curve_row is None:
            curve_row = fallback_row

        months: list[str] = []
        monthly: list[float] = []
        labels: list[str] = []
        for col_idx in range(start_col, end_col + 1):
            header_val = ws.cell(header_row, col_idx).value
            alt_header = ws.cell(1, col_idx).value
            month = parse_month_value(header_val) or parse_month_value(alt_header)
            value = safe_num(ws.cell(curve_row, col_idx).value)
            if month is None and value == 0:
                continue
            if month is None:
                continue
            months.append(month)
            monthly.append(round(value, 2))
            labels.append(month_label(month))

        cumulative: list[float] = []
        running = 0.0
        for value in monthly:
            running += value
            cumulative.append(round(running, 2))
        total = cumulative[-1] if cumulative else 0.0
        pct = [round(v / total * 100, 2) if total else 0.0 for v in cumulative]
        return {
            "months": months,
            "labels": labels,
            "monthly": monthly,
            "cumulative": cumulative,
            "pct": pct,
            "total": round(total, 2),
            "curveRow": curve_row,
        }

    def build_nodes(self, budget_records: list[dict[str, Any]], proc_records: list[dict[str, Any]]) -> dict[str, Node]:
        codes = sorted(
            {r["code"] for r in budget_records} | {r["code"] for r in proc_records},
            key=lambda x: [int(part) for part in x.split(".")],
        )
        nodes = {
            code: Node(code=code, parent=".".join(code.split(".")[:-1]) or None, depth=len(code.split(".")))
            for code in codes
        }

        for record in budget_records:
            node = nodes[record["code"]]
            node.desc_pl = first_non_empty(node.desc_pl, record["desc_pl"])
            node.desc_en = first_non_empty(node.desc_en, record["desc_en"])
            if record["contractor"]:
                node.companies.add(str(record["contractor"]))
            node.direct_budget_exec += safe_num(record["exec_budget"])
            node.direct_budget_ab += safe_num(record["budget_ab"])
            node.rows_budget += 1

        for record in proc_records:
            node = nodes[record["code"]]
            node.desc_pl = first_non_empty(node.desc_pl, record["desc_pl"])
            node.desc_en = first_non_empty(node.desc_en, record["desc_en"])
            node.buyer = first_non_empty(node.buyer, record["buyer"])
            node.site_resp = first_non_empty(node.site_resp, record["site_resp"])
            node.tme_resp = first_non_empty(node.tme_resp, record["tme_resp"])
            node.category = first_non_empty(node.category, record["category"])
            node.comment = first_non_empty(node.comment, record["comment"])
            if record["company"]:
                node.companies.add(str(record["company"]))
            if record["status"]:
                node.statuses.append(record["status"])
            node.direct_legacy_budget += safe_num(record["legacy_budget"])
            node.direct_updated_budget += safe_num(record["updated_budget"])
            node.direct_contracted += safe_num(record["contracted"])
            node.direct_result += safe_num(record["result"])
            node.rows_proc += 1
            for date_key in node.dates:
                if record[date_key]:
                    node.dates[date_key].append(record[date_key])

        for code, node in nodes.items():
            if node.parent and node.parent in nodes:
                nodes[node.parent].children.append(code)
        for node in nodes.values():
            node.children.sort(key=lambda x: [int(part) for part in x.split(".")])

        return nodes

    def should_include_direct(self, node: Node, budget_rows_by_code: dict[str, list[dict[str, Any]]], proc_rows_by_code: dict[str, list[dict[str, Any]]]) -> bool:
        for row in proc_rows_by_code.get(node.code, []):
            if any(
                row.get(key)
                for key in (
                    "status",
                    "company",
                    "buyer",
                    "site_resp",
                    "tme_resp",
                    "start_proc",
                    "deadline_proc",
                    "deadline_closing",
                    "start_exec",
                    "end_exec",
                    "updated_budget",
                    "contracted",
                )
            ):
                return True
        for row in budget_rows_by_code.get(node.code, []):
            if row.get("contractor") or row.get("order_no") or safe_num(row.get("budget_ab")):
                return True
        return False

    def rollup_nodes(self, nodes: dict[str, Node], budget_rows_by_code: dict[str, list[dict[str, Any]]], proc_rows_by_code: dict[str, list[dict[str, Any]]], today: datetime) -> None:
        completed = {normalize_status(s) for s in self.config["completed_statuses"]}

        def _walk(code: str) -> dict[str, Any]:
            node = nodes[code]
            include_direct = self.should_include_direct(node, budget_rows_by_code, proc_rows_by_code)

            budget_ab = node.direct_budget_ab if include_direct else 0.0
            budget_exec = node.direct_budget_exec if include_direct else 0.0
            legacy_budget = node.direct_legacy_budget if include_direct else 0.0
            updated_budget = node.direct_updated_budget if include_direct else 0.0
            contracted = node.direct_contracted if include_direct else 0.0
            result = node.direct_result if include_direct else 0.0
            statuses = list(node.statuses)
            deadlines = list(node.dates["deadline_closing"])
            end_dates = list(node.dates["end_exec"])

            for child_code in node.children:
                child_roll = _walk(child_code)
                budget_ab += child_roll["budget_ab"]
                budget_exec += child_roll["budget_exec"]
                legacy_budget += child_roll["legacy_budget"]
                updated_budget += child_roll["updated_budget"]
                contracted += child_roll["contracted"]
                result += child_roll["result"]
                statuses.extend(child_roll["statuses"])
                deadlines.extend(child_roll["deadlines"])
                end_dates.extend(child_roll["end_dates"])

            dominant = Counter(statuses).most_common(1)[0][0] if statuses else None
            progress = sum(STATUS_PROGRESS.get(status, 0) for status in statuses) / len(statuses) if statuses else 0.0
            min_deadline = min(deadlines) if deadlines else None
            max_end = max(end_dates) if end_dates else None
            overdue_days = 0
            if min_deadline and dominant not in completed and today.date() > min_deadline.date():
                overdue_days = (today.date() - min_deadline.date()).days
            contract_coverage = (contracted / updated_budget * 100) if updated_budget else 0.0
            budget_variance_pct = ((updated_budget - budget_ab) / budget_ab * 100) if budget_ab else 0.0
            completion_share = (sum(1 for s in statuses if s in completed) / len(statuses) * 100) if statuses else 0.0

            node.rollup = {
                "budget_ab": budget_ab,
                "budget_exec": budget_exec,
                "legacy_budget": legacy_budget,
                "updated_budget": updated_budget,
                "contracted": contracted,
                "result": result,
                "statuses": statuses,
                "dominant": dominant,
                "progress": progress,
                "completion_share": completion_share,
                "min_deadline": min_deadline,
                "max_end": max_end,
                "overdue_days": overdue_days,
                "contract_coverage": contract_coverage,
                "budget_variance_pct": budget_variance_pct,
                "deadlines": deadlines,
                "end_dates": end_dates,
            }
            return node.rollup

        roots = sorted([code for code, node in nodes.items() if not node.parent or node.parent not in nodes], key=lambda x: [int(part) for part in x.split(".")])
        for root in roots:
            _walk(root)

    def compute_health(self, node: Node) -> str:
        rr = node.rollup
        thresholds = self.config["thresholds"]
        if rr.get("overdue_days", 0) > 0:
            return "red"
        if rr.get("budget_variance_pct", 0) >= thresholds["variance_red_pct"]:
            return "red"
        if rr.get("contract_coverage", 0) <= thresholds["contract_coverage_red_pct"] and rr.get("updated_budget", 0) > 0:
            return "red"
        if rr.get("dominant") in {normalize_status(s) for s in self.config["completed_statuses"]}:
            return "green"
        if rr.get("budget_variance_pct", 0) >= thresholds["variance_amber_pct"]:
            return "amber"
        if rr.get("contract_coverage", 0) <= thresholds["contract_coverage_amber_pct"] and rr.get("updated_budget", 0) > 0:
            return "amber"
        return "green" if rr.get("updated_budget", 0) else "neutral"


    def build_curve(self, rows: list[dict[str, Any]], budget_curve: dict[str, Any], today: datetime) -> dict[str, Any]:
        proc_buckets: dict[str, float] = defaultdict(float)
        contracted_buckets: dict[str, float] = defaultdict(float)

        for row in rows:
            award_date = row.get("deadlineClosing") or row.get("deadlineProc") or row.get("startProc")
            exec_start = row.get("startExecution")
            exec_end = row.get("endExecution") or exec_start
            updated = safe_num(row.get("updatedBudget"))
            contracted = safe_num(row.get("contractedValue"))

            if award_date and updated:
                proc_buckets[award_date[:7]] += updated

            if contracted and exec_start:
                start_dt = datetime.fromisoformat(exec_start)
                end_dt = datetime.fromisoformat(exec_end) if exec_end else start_dt
                if end_dt < start_dt:
                    start_dt, end_dt = end_dt, start_dt
                span_months = iter_months(month_key(start_dt), month_key(end_dt))
                share = contracted / max(len(span_months), 1)
                for month in span_months:
                    contracted_buckets[month] += share

        all_months = set(proc_buckets) | set(contracted_buckets) | set(budget_curve.get("months", []))
        if not all_months:
            return {
                "months": [], "labels": [], "procPct": [], "budgetPct": [], "contractedPct": [],
                "procAbs": [], "budgetAbs": [], "contractedAbs": [], "procMonthly": [], "budgetMonthly": [],
                "contractedMonthly": [], "currentMarker": None, "budgetRowLabel": self.config["budget"].get("curve_label", "TKW Razem")
            }

        curve_cfg = self.config.get("curve", {})
        start_month = curve_cfg.get("start_month") or min(all_months)
        end_month = curve_cfg.get("end_month") or max(all_months)
        months = iter_months(start_month, end_month)
        budget_map = {m: v for m, v in zip(budget_curve.get("months", []), budget_curve.get("monthly", []))}
        proc_monthly: list[float] = []
        budget_monthly: list[float] = []
        contracted_monthly: list[float] = []
        proc_abs: list[float] = []
        budget_abs: list[float] = []
        contracted_abs: list[float] = []
        p = b = c = 0.0
        for month in months:
            pm = round(proc_buckets.get(month, 0.0), 2)
            bm = round(budget_map.get(month, 0.0), 2)
            cm = round(contracted_buckets.get(month, 0.0), 2)
            proc_monthly.append(pm)
            budget_monthly.append(bm)
            contracted_monthly.append(cm)
            p += pm
            b += bm
            c += cm
            proc_abs.append(round(p, 2))
            budget_abs.append(round(b, 2))
            contracted_abs.append(round(c, 2))

        p_total = proc_abs[-1] or 1.0
        b_total = budget_abs[-1] or 1.0
        c_total = contracted_abs[-1] or 1.0
        proc_pct = [round(v / p_total * 100, 2) for v in proc_abs]
        budget_pct = [round(v / b_total * 100, 2) for v in budget_abs]
        contracted_pct = [round(v / c_total * 100, 2) for v in contracted_abs]

        current_month = today.strftime("%Y-%m")
        marker_idx = 0
        for idx, month in enumerate(months):
            if month <= current_month:
                marker_idx = idx
            else:
                break

        current_marker = None
        if months:
            current_marker = {
                "index": marker_idx,
                "month": months[marker_idx],
                "label": month_label(months[marker_idx]),
                "procPct": proc_pct[marker_idx],
                "budgetPct": budget_pct[marker_idx],
                "contractedPct": contracted_pct[marker_idx],
                "procAbs": proc_abs[marker_idx],
                "budgetAbs": budget_abs[marker_idx],
                "contractedAbs": contracted_abs[marker_idx],
            }

        return {
            "months": months,
            "labels": [month_label(x) for x in months],
            "xLabelEvery": int(curve_cfg.get("x_label_every") or 2),
            "procPct": proc_pct,
            "budgetPct": budget_pct,
            "contractedPct": contracted_pct,
            "procAbs": proc_abs,
            "budgetAbs": budget_abs,
            "contractedAbs": contracted_abs,
            "procMonthly": proc_monthly,
            "budgetMonthly": budget_monthly,
            "contractedMonthly": contracted_monthly,
            "currentMarker": current_marker,
            "budgetRowLabel": self.config["budget"].get("curve_label", "TKW Razem"),
        }

    def build_timeline(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        items = []
        for row in rows:
            start = row.get("startProc") or row.get("deadlineProc") or row.get("deadlineClosing") or row.get("startExecution")
            end = row.get("deadlineClosing") or row.get("startExecution") or row.get("endExecution") or start
            if not start and not end:
                continue
            if start and end and start > end:
                start, end = end, start
            items.append({
                "code": row["code"],
                "name": row["name"],
                "status": row["status"],
                "tone": tone_from_status(row["status"]),
                "start": start,
                "end": end,
                "updatedBudget": row["updatedBudget"],
                "overdueDays": row["overdueDays"],
            })
        if not items:
            return {"months": [], "labels": [], "items": []}
        items = sorted(items, key=lambda x: ((x["end"] or "9999-12-31"), -(x["updatedBudget"] or 0), x["code"]))[:10]
        all_months = []
        for item in items:
            if item["start"]:
                all_months.append(item["start"][:7])
            if item["end"]:
                all_months.append(item["end"][:7])
        months = iter_months(min(all_months), max(all_months)) if all_months else []
        return {"months": months, "labels": [month_label(x) for x in months], "items": items}

    def read_ref_values(self, workbook, ref: str) -> list[Any]:
        sheet_name, cells = parse_sheet_range(ref)
        ws = workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(cells.replace('$', ''))
        values: list[Any] = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            values.extend(list(row))
        return values

    def extract_portfolio_curve(self, curve_path: Path | None) -> dict[str, Any]:
        if not curve_path or not curve_path.exists():
            return {"title": None, "labels": [], "series": [], "sourceSheet": self.config.get("portfolio_curve", {}).get("sheet_name")}
        wb = openpyxl.load_workbook(curve_path, data_only=True)
        sheet_name = self.config.get("portfolio_curve", {}).get("sheet_name", "10288_ITPO PROJEKT-SCuvrve")
        if sheet_name not in wb.sheetnames:
            return {"title": None, "labels": [], "series": [], "sourceSheet": sheet_name}
        ws = wb[sheet_name]
        charts = getattr(ws, '_charts', []) or []
        if not charts:
            return {"title": None, "labels": [], "series": [], "sourceSheet": sheet_name}
        chart = charts[0]
        labels: list[str] = []
        series_out: list[dict[str, Any]] = []
        palette = ['#2f68c8', '#7EA73B', '#9AA5B5', '#D59C1A']
        for idx, ser in enumerate(chart.ser):
            title = getattr(getattr(ser, 'tx', None), 'v', None) or f'Series {idx+1}'
            cat_ref_obj = getattr(ser, 'cat', None)
            cat_ref = None
            if cat_ref_obj is not None:
                cat_ref_obj = getattr(cat_ref_obj, 'numRef', None) or getattr(cat_ref_obj, 'strRef', None)
                cat_ref = getattr(cat_ref_obj, 'f', None)
            val_ref_obj = getattr(getattr(ser, 'val', None), 'numRef', None)
            val_ref = getattr(val_ref_obj, 'f', None)
            if cat_ref and not labels:
                raw_labels = self.read_ref_values(wb, cat_ref)
                labels = [normalize_chart_label(v) for v in raw_labels if v not in (None, '')]
            if not val_ref:
                continue
            raw_vals = self.read_ref_values(wb, val_ref)
            values = [to_float_or_none(v) for v in raw_vals]
            max_abs = max((abs(v) for v in values if v is not None), default=0)
            if max_abs <= 1.2:
                values = [None if v is None else round(v * 100, 2) for v in values]
            else:
                values = [None if v is None else round(v, 2) for v in values]
            if labels and len(values) < len(labels):
                values.extend([None] * (len(labels) - len(values)))
            elif labels and len(values) > len(labels):
                values = values[:len(labels)]
            series_out.append({"name": title, "values": values, "color": palette[idx % len(palette)]})
        return {"title": "Project S-Curve", "labels": labels, "series": series_out, "sourceSheet": sheet_name}


    def extract_ec_decision(self, ec_path: Path | None) -> dict[str, Any]:
        cfg = self.config.get("ec_decision", {})
        sheet_name = cfg.get("sheet_name", "EC decision")
        if not ec_path or not ec_path.exists():
            return {"title": "EC decision", "headers": [], "rows": [], "sourceSheet": sheet_name}
        wb = openpyxl.load_workbook(ec_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {"title": "EC decision", "headers": [], "rows": [], "sourceSheet": sheet_name}
        ws = wb[sheet_name]
        cols = cfg.get("columns", ["D", "E", "G", "I", "J", "K", "L", "M"])
        header_row = int(cfg.get("header_row", 3))
        data_start = int(cfg.get("data_start_row", 4))
        headers = []
        for col in cols:
            val = ws[f"{col}{header_row}"].value
            headers.append(str(val).strip() if val not in (None, "") else col)
        rows = []
        for r in range(data_start, ws.max_row + 1):
            vals = []
            nonempty = False
            for col in cols:
                v = ws[f"{col}{r}"].value
                if isinstance(v, datetime):
                    v = v.date().isoformat()
                elif isinstance(v, date):
                    v = v.isoformat()
                elif isinstance(v, float):
                    v = round(v, 2)
                if v not in (None, ""):
                    nonempty = True
                vals.append(v)
            if nonempty:
                rows.append(vals)
        return {"title": "EC decision", "headers": headers, "rows": rows, "sourceSheet": sheet_name}

    def extract_status_progress(self, status_path: Path | None) -> dict[str, Any]:
        cfg = self.config.get("status_progress", {})
        sheet_name = cfg.get("sheet_name", "Eng_Proc_Cons")
        if not status_path or not status_path.exists():
            return {"title": "Status progress", "headers": [], "rows": [], "sourceSheet": sheet_name}
        wb = openpyxl.load_workbook(status_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {"title": "Status progress", "headers": [], "rows": [], "sourceSheet": sheet_name}
        ws = wb[sheet_name]
        cols = cfg.get("columns", ["A", "B", "C"])
        header_row = int(cfg.get("header_row", 1))
        data_start = int(cfg.get("data_start_row", 2))
        headers = []
        for col in cols:
            val = ws[f"{col}{header_row}"].value
            headers.append(str(val).strip() if val not in (None, "") else col)
        rows = []
        for r in range(data_start, ws.max_row + 1):
            vals = []
            nonempty = False
            for i, col in enumerate(cols):
                v = ws[f"{col}{r}"].value
                if i > 0 and v not in (None, ""):
                    try:
                        v = round(float(v) * 100 if abs(float(v)) <= 1.2 else float(v), 1)
                    except Exception:
                        pass
                if v not in (None, ""):
                    nonempty = True
                vals.append(v)
            if nonempty:
                rows.append(vals)
        return {"title": "Status progress", "headers": headers, "rows": rows, "sourceSheet": sheet_name}

    def build_payload(
        self,
        budget_summary: dict[str, Any],
        proc_summary: dict[str, Any],
        budget_records: list[dict[str, Any]],
        proc_records: list[dict[str, Any]],
        nodes: dict[str, Node],
        budget_path: Path,
        proc_path: Path,
        today: datetime,
    ) -> dict[str, Any]:
        completed = {normalize_status(s) for s in self.config["completed_statuses"]}
        contract_prep_status = normalize_status(self.config["contract_prep_status"])

        weighted_progress_numerator = sum(STATUS_PROGRESS.get(r["status"], 0) * safe_num(r["updated_budget"]) for r in proc_records if r["status"])
        weighted_progress_denominator = sum(safe_num(r["updated_budget"]) for r in proc_records if r["status"])
        weighted_progress_pct = weighted_progress_numerator / weighted_progress_denominator if weighted_progress_denominator else 0.0

        completion_value = sum(safe_num(r["updated_budget"]) for r in proc_records if r["status"] in completed)
        completion_pct = completion_value / proc_summary["updatedBudgetTotal"] * 100 if proc_summary["updatedBudgetTotal"] else 0.0
        contracted_coverage_pct = proc_summary["contractedTotal"] / proc_summary["updatedBudgetTotal"] * 100 if proc_summary["updatedBudgetTotal"] else 0.0
        variance_amount = proc_summary["updatedBudgetTotal"] - budget_summary["budgetAbTotal"]
        variance_pct = variance_amount / budget_summary["budgetAbTotal"] * 100 if budget_summary["budgetAbTotal"] else 0.0

        counts_by_status = Counter(r["status"] for r in proc_records if r["status"])
        values_by_status = defaultdict(float)
        for r in proc_records:
            if r["status"]:
                values_by_status[r["status"]] += safe_num(r["updated_budget"]) if safe_num(r["updated_budget"]) else safe_num(r["contracted"])

        status_funnel = []
        for status in STATUS_ORDER:
            if counts_by_status.get(status, 0):
                status_funnel.append(
                    {
                        "status": status,
                        "count": counts_by_status[status],
                        "value": round(values_by_status[status], 2),
                        "progress": STATUS_PROGRESS.get(status, 0),
                        "tone": tone_from_status(status),
                    }
                )

        direct_rows: list[dict[str, Any]] = []
        for code, node in nodes.items():
            has_direct = bool(
                node.direct_updated_budget
                or node.direct_contracted
                or node.direct_budget_ab
                or node.statuses
                or node.companies
                or any(node.dates[key] for key in node.dates)
            )
            if not has_direct:
                continue
            status = Counter(node.statuses).most_common(1)[0][0] if node.statuses else None
            deadline = min(node.dates["deadline_closing"]) if node.dates["deadline_closing"] else None
            start_proc = min(node.dates["start_proc"]) if node.dates["start_proc"] else None
            deadline_proc = min(node.dates["deadline_proc"]) if node.dates["deadline_proc"] else None
            start_exec = min(node.dates["start_exec"]) if node.dates["start_exec"] else None
            end_exec = max(node.dates["end_exec"]) if node.dates["end_exec"] else None
            overdue_days = 0
            if deadline and status not in completed and today.date() > deadline.date():
                overdue_days = (today.date() - deadline.date()).days
            variance_direct = node.direct_updated_budget - node.direct_budget_ab
            variance_direct_pct = (variance_direct / node.direct_budget_ab * 100) if node.direct_budget_ab else 0.0
            root_code = code.split(".")[0]
            direct_rows.append(
                {
                    "code": code,
                    "rootCode": root_code,
                    "depth": node.depth,
                    "name": node.desc_en or node.desc_pl or code,
                    "buyer": node.buyer,
                    "siteResp": node.site_resp,
                    "tmeResp": node.tme_resp,
                    "category": node.category,
                    "status": status,
                    "statusTone": tone_from_status(status),
                    "budgetAb": round(node.direct_budget_ab, 2),
                    "updatedBudget": round(node.direct_updated_budget, 2),
                    "legacyBudget": round(node.direct_legacy_budget, 2),
                    "contractedValue": round(node.direct_contracted, 2),
                    "result": round(node.direct_result, 2),
                    "deadlineClosing": to_iso(deadline),
                    "startProc": to_iso(start_proc),
                    "deadlineProc": to_iso(deadline_proc),
                    "startExecution": to_iso(start_exec),
                    "endExecution": to_iso(end_exec),
                    "company": ", ".join(sorted(node.companies)) if node.companies else None,
                    "comment": node.comment,
                    "overdueDays": overdue_days,
                    "varianceAmount": round(variance_direct, 2),
                    "variancePct": round(variance_direct_pct, 2),
                    "progress": STATUS_PROGRESS.get(status, 0),
                }
            )

        open_rows = [row for row in direct_rows if row["status"] not in completed]
        contract_prep_rows = [row for row in direct_rows if row["status"] == contract_prep_status]
        pct_rows = [row for row in direct_rows if row["status"] == "PCT/approval"]
        specs_rows = [row for row in direct_rows if row["status"] in {"enquiry", "negotiation"}]
        overdue_rows = [row for row in open_rows if row["overdueDays"] > 0]

        def critical_score(item: dict[str, Any]) -> float:
            value_weight = item["updatedBudget"] / max(proc_summary["updatedBudgetTotal"], 1) * 100
            delay_weight = min(item["overdueDays"], 120) * 0.8
            contract_gap = max(item["updatedBudget"] - item["contractedValue"], 0) / max(item["updatedBudget"], 1) * 25 if item["updatedBudget"] else 0
            return delay_weight + value_weight + contract_gap

        critical_packages = sorted(open_rows, key=critical_score, reverse=True)[:10]
        top_overruns = sorted([row for row in direct_rows if row["varianceAmount"] > 0], key=lambda x: x["varianceAmount"], reverse=True)[:10]

        root_groups = []
        delayed_by_root = []
        for root_code in sorted([code for code in nodes if code.split(".")[0] == code], key=lambda x: [int(part) for part in x.split(".")]):
            node = nodes[root_code]
            rr = node.rollup
            root_groups.append(
                {
                    "code": root_code,
                    "name": node.desc_en or node.desc_pl or root_code,
                    "health": self.compute_health(node),
                    "updatedBudget": round(rr["updated_budget"], 2),
                    "contractedValue": round(rr["contracted"], 2),
                    "budgetAb": round(rr["budget_ab"], 2),
                    "completionPct": round(rr["completion_share"], 1),
                    "contractCoveragePct": round(rr["contract_coverage"], 1),
                    "variancePct": round(rr["budget_variance_pct"], 1),
                    "overdueDays": rr["overdue_days"],
                    "dominantStatus": rr["dominant"],
                }
            )
            root_open = [r for r in open_rows if r["rootCode"] == root_code]
            delayed_by_root.append(
                {
                    "code": root_code,
                    "name": node.desc_en or node.desc_pl or root_code,
                    "count": len([r for r in root_open if r["overdueDays"] > 0]),
                    "value": round(sum(r["updatedBudget"] for r in root_open if r["overdueDays"] > 0), 2),
                }
            )
        root_groups = sorted(root_groups, key=lambda x: x["updatedBudget"], reverse=True)
        delayed_by_root = sorted(delayed_by_root, key=lambda x: (x["value"], x["count"]), reverse=True)[:8]

        upcoming = sorted(
            [row for row in open_rows if row["deadlineClosing"]],
            key=lambda x: (x["deadlineClosing"], -x["updatedBudget"]),
        )[:10]

        category_groups: dict[str, dict[str, Any]] = defaultdict(lambda: {"name": "Uncategorized", "count": 0, "updatedBudget": 0.0, "contractedValue": 0.0})
        for row in direct_rows:
            key = row["category"] or "Uncategorized"
            category_groups[key]["name"] = key
            category_groups[key]["count"] += 1
            category_groups[key]["updatedBudget"] += row["updatedBudget"]
            category_groups[key]["contractedValue"] += row["contractedValue"]
        category_breakdown = sorted(category_groups.values(), key=lambda x: x["updatedBudget"], reverse=True)[:8]
        for item in category_breakdown:
            item["updatedBudget"] = round(item["updatedBudget"], 2)
            item["contractedValue"] = round(item["contractedValue"], 2)
            item["sharePct"] = round(item["updatedBudget"] / max(proc_summary["updatedBudgetTotal"], 1) * 100, 2)

        total_updated = max(proc_summary["updatedBudgetTotal"], 1)
        cost_breakdown = []
        for rg in root_groups[:8]:
            cost_breakdown.append({
                "code": rg["code"],
                "name": rg["name"],
                "updatedBudget": rg["updatedBudget"],
                "budgetAb": rg["budgetAb"],
                "sharePct": round(rg["updatedBudget"] / total_updated * 100, 2),
                "health": rg["health"],
            })

        orders_closing_rows = sorted(contract_prep_rows + pct_rows, key=lambda x: (x["status"] != "PCT/approval", -x["updatedBudget"] ))[:5]
        specs_rows_sorted = sorted(specs_rows, key=lambda x: (-x["updatedBudget"], x["code"]))[:5]
        budget_curve = self.extract_budget_curve(budget_path)
        s_curve = self.build_curve(direct_rows, budget_curve, today)
        timeline = self.build_timeline(upcoming if upcoming else open_rows)

        order_mix = [
            {
                "label": "Ordinato",
                "count": len([r for r in direct_rows if r["status"] in completed]),
                "value": round(sum(r["updatedBudget"] for r in direct_rows if r["status"] in completed), 2),
                "tone": "green",
            },
            {
                "label": "In negoziazione",
                "count": len([r for r in direct_rows if r["status"] in {"negotiation", contract_prep_status, "PCT/approval"}]),
                "value": round(sum(r["updatedBudget"] for r in direct_rows if r["status"] in {"negotiation", contract_prep_status, "PCT/approval"}), 2),
                "tone": "amber",
            },
            {
                "label": "Da ordinare",
                "count": len([r for r in direct_rows if r["status"] in {"planned", "enquiry"} or not r["status"]]),
                "value": round(sum(r["updatedBudget"] for r in direct_rows if r["status"] in {"planned", "enquiry"} or not r["status"]), 2),
                "tone": "red",
            },
        ]
        order_mix_total = sum(x["count"] for x in order_mix) or 1
        for item in order_mix:
            item["sharePct"] = round(item["count"] / order_mix_total * 100, 1)

        if overdue_rows or variance_pct >= self.config["thresholds"]["variance_red_pct"] or contracted_coverage_pct <= self.config["thresholds"]["contract_coverage_red_pct"]:
            overall = "AT RISK"
            overall_tone = "amber"
        else:
            overall = "ON TRACK"
            overall_tone = "green"

        executive_message = self.config["executive_note"]
        if overdue_rows:
            biggest = sorted(overdue_rows, key=lambda x: (x["overdueDays"], x["updatedBudget"]), reverse=True)[0]
            executive_message = f"Main risk today: {biggest['code']} {biggest['name']} is overdue by {biggest['overdueDays']} days with {biggest['updatedBudget']:.0f} PLN still exposed."
        elif variance_pct > 0:
            executive_message = f"Updated procurement budget is {variance_pct:.1f}% above the AB baseline."

        snapshots = {
            "closedCount": len([r for r in direct_rows if r["status"] in completed]),
            "openCount": len(open_rows),
            "overdueCount": len(overdue_rows),
            "contractPrepCount": len(contract_prep_rows),
            "pctApprovalCount": len(pct_rows),
            "contractPrepValue": round(sum(r["updatedBudget"] for r in contract_prep_rows), 2),
            "pctApprovalValue": round(sum(r["updatedBudget"] for r in pct_rows), 2),
            "specsIssuedCount": len(specs_rows),
            "specsIssuedValue": round(sum(r["updatedBudget"] for r in specs_rows), 2),
        }

        return {
            "meta": {
                "projectTitle": self.config["project_title"],
                "generatedAt": today.isoformat(timespec="seconds"),
                "currency": self.config["currency"],
                "sourceFiles": {
                    "budget": budget_path.name,
                    "procurement": proc_path.name,
                    "scurve": None,
                },
                "columnMapping": {
                    "budgetTotal": self.config["budget"]["total_column"],
                    "updatedBudget": self.config["procurement"]["updated_budget_column"],
                    "contracted": self.config["procurement"]["contracted_column"],
                },
            },
            "summary": {
                "overallStatus": overall,
                "overallTone": overall_tone,
                "budgetAbTotal": round(budget_summary["budgetAbTotal"], 2),
                "updatedBudgetTotal": round(proc_summary["updatedBudgetTotal"], 2),
                "contractedTotal": round(proc_summary["contractedTotal"], 2),
                "legacyBudgetTotal": round(proc_summary["legacyBudgetTotal"], 2),
                "varianceAmount": round(variance_amount, 2),
                "variancePct": round(variance_pct, 2),
                "contractCoveragePct": round(contracted_coverage_pct, 2),
                "completionPct": round(completion_pct, 2),
                "weightedProgressPct": round(weighted_progress_pct, 2),
                "valueToAward": round(proc_summary["updatedBudgetTotal"] - proc_summary["contractedTotal"], 2),
                "curveCurrentMonth": s_curve.get("currentMarker", {}).get("label") if s_curve.get("currentMarker") else None,
                "curveCurrentProc": s_curve.get("currentMarker", {}).get("procAbs") if s_curve.get("currentMarker") else 0,
                "curveCurrentBudget": s_curve.get("currentMarker", {}).get("budgetAbs") if s_curve.get("currentMarker") else 0,
                "executiveMessage": executive_message,
                **snapshots,
            },
            "overview": {
                "sCurve": s_curve,
                "costBreakdown": cost_breakdown,
                "categoryBreakdown": category_breakdown,
                "ordersClosing": {
                    "count": len(contract_prep_rows) + len(pct_rows),
                    "value": round(sum(r["updatedBudget"] for r in contract_prep_rows + pct_rows), 2),
                    "items": orders_closing_rows,
                },
                "specsIssued": {
                    "count": len(specs_rows),
                    "value": round(sum(r["updatedBudget"] for r in specs_rows), 2),
                    "items": specs_rows_sorted,
                },
                "orderMix": order_mix,
                "budgetCurveMeta": {
                    "label": self.config["budget"].get("curve_label", "TKW Razem"),
                    "row": budget_curve.get("curveRow"),
                    "total": budget_curve.get("total", 0),
                },
                "timeline": timeline,
            },
            "statusFunnel": status_funnel,
            "criticalPackages": critical_packages,
            "topOverruns": top_overruns,
            "rootGroups": root_groups,
            "delayedByRoot": delayed_by_root,
            "upcomingMilestones": upcoming,
            "directPackages": direct_rows,
            "notes": [
                f"Total budget KPI comes from Budget column {self.config['budget']['total_column']} row {self.config['budget']['summary_row']}.",
                f"Updated budget comes from Procurement column {self.config['procurement']['updated_budget_column']} row {self.config['procurement']['summary_row']}.",
                f"Contracted value comes from Procurement column {self.config['procurement']['contracted_column']} row {self.config['procurement']['summary_row']}.",
                f"S-curve green line comes from budget row '{self.config['budget'].get('curve_label', 'TKW Razem')}' distributed on monthly columns.",
                "New WBS child rows are accepted automatically as long as the item code follows the numeric hierarchy.",
            ],
        }

    def generate_payload(self, budget_path: Path, proc_path: Path, curve_path: Path | None = None, ec_path: Path | None = None, status_path: Path | None = None) -> dict[str, Any]:
        today = datetime.now()
        budget_records, budget_summary = self.read_budget_records(budget_path)
        proc_records, proc_summary = self.read_proc_records(proc_path)
        nodes = self.build_nodes(budget_records, proc_records)
        budget_rows_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
        proc_rows_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in budget_records:
            budget_rows_by_code[row["code"]].append(row)
        for row in proc_records:
            proc_rows_by_code[row["code"]].append(row)
        self.rollup_nodes(nodes, budget_rows_by_code, proc_rows_by_code, today)
        payload = self.build_payload(budget_summary, proc_summary, budget_records, proc_records, nodes, budget_path, proc_path, today)
        payload["portfolioCurve"] = self.extract_portfolio_curve(curve_path)
        payload["ecDecision"] = self.extract_ec_decision(ec_path)
        payload["statusProgress"] = self.extract_status_progress(status_path)
        if curve_path:
            payload.setdefault("meta", {}).setdefault("sourceFiles", {})["scurve"] = curve_path.name
            try:
                mtime = datetime.fromtimestamp(curve_path.stat().st_mtime)
                payload.setdefault("meta", {})["scurveReferenceMonth"] = month_label(mtime.strftime("%Y-%m"))
                payload.setdefault("meta", {})["scurveUploadedAt"] = mtime.isoformat(timespec="seconds")
            except Exception:
                pass
        if ec_path:
            payload.setdefault("meta", {}).setdefault("sourceFiles", {})["ecdecision"] = ec_path.name
        if status_path:
            payload.setdefault("meta", {}).setdefault("sourceFiles", {})["statusprogress"] = status_path.name
        return payload


def save_snapshot(payload: dict[str, Any], output_path: Path, snapshots_dir: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    snapshots_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    snapshot_path = snapshots_dir / f"dashboard-{stamp}.json"
    shutil.copy2(output_path, snapshot_path)
    return snapshot_path
