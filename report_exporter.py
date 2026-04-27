from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
SUB_FILL = PatternFill('solid', fgColor='DCE6F1')
ACCENT_FILL = PatternFill('solid', fgColor='F4B183')
THIN = Side(style='thin', color='D9E2F3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fmt_pln(x, _pos=None):
    return f"{x:,.0f} PLN".replace(',', '.')


def _safe_name(text: str) -> str:
    return ''.join(ch if ch.isalnum() or ch in '-_.' else '_' for ch in text)


def _write_table(ws, start_row: int, start_col: int, headers: list[str], rows: list[list[Any]], title: str | None = None) -> int:
    row = start_row
    if title:
        ws.cell(row, start_col, title)
        ws.cell(row, start_col).font = Font(bold=True, size=13, color='1F1F1F')
        row += 1
    for c, header in enumerate(headers, start_col):
        cell = ws.cell(row, c, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    row += 1
    for values in rows:
        for c, value in enumerate(values, start_col):
            cell = ws.cell(row, c, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top')
        row += 1
    return row


def _autofit(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = 0
        for cell in col_cells:
            val = '' if cell.value is None else str(cell.value)
            width = max(width, min(len(val) + 2, max_width))
        ws.column_dimensions[letter].width = max(min_width, width)


def _plot_s_curve(curve: dict[str, Any], out_path: Path) -> None:
    labels = curve.get('labels', [])
    if not labels:
        fig, ax = plt.subplots(figsize=(10, 4.8))
        ax.text(0.5, 0.5, 'Nessun dato curva disponibile', ha='center', va='center')
        ax.axis('off')
        fig.savefig(out_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return
    x = list(range(len(labels)))
    proc_abs = curve.get('procAbs', [])
    budget_abs = curve.get('budgetAbs', [])
    marker = curve.get('currentMarker') or {}

    fig, ax = plt.subplots(figsize=(11.8, 5.6))
    ax.plot(x, proc_abs, color='#2F68C8', linewidth=2.8, label='Procurement schedule cumulativo')
    ax.plot(x, budget_abs, color='#7EA73B', linewidth=2.8, label='Budget / Programma cliente cumulativo')
    if marker:
        idx = marker.get('index', 0)
        ax.scatter([idx], [proc_abs[idx]], s=52, color='#C61E1E', zorder=4, label='Stato attuale')
        ax.axvline(idx, color='#C61E1E', linestyle=':', linewidth=1.0, alpha=0.5)
    ax.set_title('Andamento degli ordini / Curva S cumulata', fontsize=15, weight='bold', pad=16)
    ax.yaxis.set_major_formatter(FuncFormatter(_fmt_pln))
    ax.grid(axis='y', color='#D9E2F3', linewidth=0.8)
    ax.set_axisbelow(True)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=90, fontsize=8)
    ax.tick_params(axis='y', labelsize=8)
    ax.legend(loc='upper left', frameon=False, ncol=2, fontsize=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    fig.savefig(out_path, dpi=160, bbox_inches='tight')
    plt.close(fig)


def _plot_order_mix(order_mix: list[dict[str, Any]], out_path: Path) -> None:
    labels = [f"{item.get('label')}\n{item.get('count',0)}; {item.get('sharePct',0):.0f}%" for item in order_mix]
    sizes = [max(float(item.get('count', 0)), 0.0001) for item in order_mix]
    colors = []
    tone_map = {'green': '#2F9D58', 'amber': '#D59C1A', 'red': '#A61515', 'blue': '#2F68C8', 'neutral': '#B0B7C3'}
    for item in order_mix:
        colors.append(tone_map.get(item.get('tone'), '#B0B7C3'))

    fig, ax = plt.subplots(figsize=(6.5, 4.6))
    wedges, texts = ax.pie(
        sizes,
        labels=labels,
        colors=colors,
        startangle=35,
        labeldistance=1.15,
        wedgeprops={'linewidth': 1, 'edgecolor': 'white'},
        textprops={'fontsize': 8}
    )
    ax.set_title('Mix ordini', fontsize=13, weight='bold', pad=12)
    ax.axis('equal')
    fig.tight_layout()
    fig.savefig(out_path, dpi=160, bbox_inches='tight')
    plt.close(fig)


def generate_excel_report(payload: dict[str, Any], reports_dir: Path) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    file_name = f"Combined_S_Curve_Procurement_{timestamp}.xlsx"
    report_path = reports_dir / file_name

    tmp_curve = reports_dir / f"_curve_{timestamp}.png"
    tmp_pie = reports_dir / f"_pie_{timestamp}.png"

    summary = payload.get('summary', {})
    overview = payload.get('overview', {})
    meta = payload.get('meta', {})
    curve = overview.get('sCurve', {})
    order_mix = overview.get('orderMix', [])

    _plot_s_curve(curve, tmp_curve)
    _plot_order_mix(order_mix, tmp_pie)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Executive Summary'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A6'

    ws['A1'] = meta.get('projectTitle', 'Waste-to-Energy Dashboard')
    ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Report combinato Curva S + Procurement Schedule'
    ws['A2'].font = Font(size=12, bold=True)
    ws.merge_cells('A2:H2')
    ws['A3'] = f"Generato: {meta.get('generatedAt', '')}"
    ws['D3'] = f"Budget file: {meta.get('sourceFiles', {}).get('budget', '—')}"
    ws['G3'] = f"Procurement file: {meta.get('sourceFiles', {}).get('procurement', '—')}"

    kpis = [
        ('Overall status', summary.get('overallStatus', '—')),
        ('Updated Budget', summary.get('updatedBudgetTotal', 0)),
        ('Baseline Budget', summary.get('budgetAbTotal', 0)),
        ('Contracted', summary.get('contractedTotal', 0)),
        ('Completion %', summary.get('completionPct', 0)),
        ('Variance %', summary.get('variancePct', 0)),
        ('Contract prep. count', summary.get('contractPrepCount', 0)),
        ('Packages overdue', summary.get('overdueCount', 0)),
    ]
    row = 5
    for idx, (label, value) in enumerate(kpis, start=0):
        r = row + idx // 4 * 2
        c = 1 + (idx % 4) * 2
        ws.cell(r, c, label)
        ws.cell(r, c).fill = SUB_FILL
        ws.cell(r, c).font = Font(bold=True)
        ws.cell(r, c).border = BORDER
        ws.cell(r, c+1, value)
        ws.cell(r, c+1).border = BORDER
        if 'Budget' in label or label == 'Contracted':
            ws.cell(r, c+1).number_format = '#,##0'
        if '%' in label:
            ws.cell(r, c+1).number_format = '0.0'

    ws['A10'] = 'Executive note'
    ws['A10'].fill = ACCENT_FILL
    ws['A10'].font = Font(bold=True)
    ws['A11'] = summary.get('executiveMessage', '')
    ws['A11'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A11:H12')

    ws.add_image(XLImage(str(tmp_curve)), 'A14')
    ws.add_image(XLImage(str(tmp_pie)), 'J15')

    # Data sheet
    ds = wb.create_sheet('S_Curve_Data')
    ds.freeze_panes = 'A2'
    headers = ['Month', 'Label', 'Proc Monthly', 'Proc Cumulative', 'Proc %', 'Budget Monthly', 'Budget Cumulative', 'Budget %', 'Current Marker']
    rows = []
    marker_idx = (curve.get('currentMarker') or {}).get('index')
    for i, month in enumerate(curve.get('months', [])):
        rows.append([
            month,
            curve.get('labels', [])[i],
            curve.get('procMonthly', [0]*len(curve.get('months', [])))[i],
            curve.get('procAbs', [0]*len(curve.get('months', [])))[i],
            curve.get('procPct', [0]*len(curve.get('months', [])))[i],
            curve.get('budgetMonthly', [0]*len(curve.get('months', [])))[i],
            curve.get('budgetAbs', [0]*len(curve.get('months', [])))[i],
            curve.get('budgetPct', [0]*len(curve.get('months', [])))[i],
            'Current' if marker_idx == i else '',
        ])
    _write_table(ds, 1, 1, headers, rows)
    for col in ['C','D','F','G']:
        for cell in ds[col][1:]:
            cell.number_format = '#,##0'
    for col in ['E','H']:
        for cell in ds[col][1:]:
            cell.number_format = '0.0'

    os = wb.create_sheet('Orders_Status')
    mix_rows = [[item.get('label'), item.get('count'), item.get('value'), item.get('sharePct')] for item in order_mix]
    next_row = _write_table(os, 1, 1, ['Bucket', 'Count', 'Value', 'Share %'], mix_rows, 'Order mix')
    next_row += 1
    closing_rows = [[x.get('code'), x.get('name'), x.get('status'), x.get('updatedBudget'), x.get('deadlineClosing') or x.get('deadlineProc') or x.get('startProc')] for x in overview.get('ordersClosing', {}).get('items', [])]
    next_row = _write_table(os, next_row, 1, ['WBS', 'Package', 'Status', 'Updated Budget', 'Date'], closing_rows, 'Ordini in chiusura')
    next_row += 1
    specs_rows = [[x.get('code'), x.get('name'), x.get('status'), x.get('updatedBudget'), x.get('deadlineClosing') or x.get('deadlineProc') or x.get('startProc')] for x in overview.get('specsIssued', {}).get('items', [])]
    _write_table(os, next_row, 1, ['WBS', 'Package', 'Status', 'Updated Budget', 'Date'], specs_rows, 'Specifiche emesse')

    dp = wb.create_sheet('Direct_Packages')
    package_rows = [[
        x.get('code'), x.get('name'), x.get('status'), x.get('buyer'), x.get('category'), x.get('budgetAb'), x.get('updatedBudget'), x.get('contractedValue'), x.get('varianceAmount'), x.get('deadlineClosing'), x.get('overdueDays')
    ] for x in payload.get('directPackages', [])[:300]]
    _write_table(dp, 1, 1, ['WBS', 'Package', 'Status', 'Buyer', 'Category', 'Baseline Budget', 'Updated Budget', 'Contracted', 'Variance', 'Deadline Closing', 'Overdue Days'], package_rows)

    notes = wb.create_sheet('Notes')
    notes['A1'] = 'Note e mapping'
    notes['A1'].font = Font(size=14, bold=True)
    for i, note in enumerate(payload.get('notes', []), start=3):
        notes[f'A{i}'] = note

    for sheet in wb.worksheets:
        _autofit(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1 and cell.column == 1 and sheet.title != 'Executive Summary':
                    pass

    wb.save(report_path)

    for path in [tmp_curve, tmp_pie]:
        try:
            path.unlink(missing_ok=True)
        except Exception:
            pass
    return report_path
